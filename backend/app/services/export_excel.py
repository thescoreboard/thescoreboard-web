"""
Excel export — builds a tournament data workbook (.xlsx) for organisers.

Deliberately takes plain dicts/lists rather than SQLAlchemy models: keeps
workbook generation isolated from the DB session and the live-scoring path
(matches.py / ws/), so it can't be accidentally wired into anything
request-per-score-update.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="FF2A2A2A", end_color="FF2A2A2A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
TITLE_FONT = Font(bold=True, size=13)


def _write_sheet(ws, headers: list[str], rows: list[list], number_cols: set[int] = frozenset()):
    """Bold header row, freeze it, write rows, auto-size columns.
    number_cols is a 0-indexed set of columns that get a plain integer format."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(row)

    widths = [len(str(h)) for h in headers]
    for row in rows:
        for i, val in enumerate(row):
            widths[i] = max(widths[i], len(str(val)) if val is not None else 0)
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(max(w + 2, 10), 45)

    for col in number_cols:
        letter = get_column_letter(col + 1)
        for cell in ws[letter][1:]:
            cell.number_format = "0"


def build_tournament_workbook(
    tournament: dict,
    event_summaries: list[dict],
    standings_rows: list[dict],
    match_rows: list[dict],
    participant_rows: list[dict],
) -> io.BytesIO:
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = tournament["name"]
    ws["A1"].font = TITLE_FONT
    meta_rows = [
        ("Venue", ", ".join(filter(None, [tournament.get("venue"), tournament.get("city"), tournament.get("state")])) or "—"),
        ("Dates", " to ".join(filter(None, [tournament.get("start_date"), tournament.get("end_date")])) or "—"),
        ("Status", (tournament.get("status") or "").title()),
        ("Total Events", len(event_summaries)),
        ("Total Participants", sum(e["participant_count"] for e in event_summaries)),
        ("Total Matches", sum(e["match_count"] for e in event_summaries)),
    ]
    for i, (label, value) in enumerate(meta_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    table_start = len(meta_rows) + 5
    ws.cell(row=table_start - 1, column=1, value="Events").font = Font(bold=True, size=12)
    headers = ["Event", "Sport", "Format", "Type", "Participants", "Matches", "Completed", "Status"]
    for i, h in enumerate(headers):
        c = ws.cell(row=table_start, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for r, e in enumerate(event_summaries, start=table_start + 1):
        ws.cell(row=r, column=1, value=e["name"])
        ws.cell(row=r, column=2, value=e["sport_key"])
        ws.cell(row=r, column=3, value=(e["format"] or "").replace("_", " ").title() or "—")
        ws.cell(row=r, column=4, value=e["participant_type"].title())
        ws.cell(row=r, column=5, value=e["participant_count"])
        ws.cell(row=r, column=6, value=e["match_count"])
        ws.cell(row=r, column=7, value=e["done_count"])
        ws.cell(row=r, column=8, value=(e["status"] or "").title())
    for col, w in zip("ABCDEFGH", [24, 14, 16, 10, 12, 10, 11, 11]):
        ws.column_dimensions[col].width = w

    # ── Sheet 2: Standings (only if any round-robin / group-stage data exists) ──
    if standings_rows:
        ws2 = wb.create_sheet("Standings")
        _write_sheet(
            ws2,
            ["Event", "Group", "Rank", "Name", "Played", "Wins", "Losses", "Points For", "Points Against", "Ranking Points"],
            [
                [r["event_name"], r["group_name"], r["rank"], r["participant_name"],
                 r["matches_played"], r["wins"], r["losses"], r["points_for"], r["points_against"], r["ranking_points"]]
                for r in standings_rows
            ],
            number_cols={2, 4, 5, 6, 7, 8, 9},
        )

    # ── Sheet 3: Match Results ───────────────────────────────────
    ws3 = wb.create_sheet("Match Results")
    _write_sheet(
        ws3,
        ["Event", "Sport", "Stage", "Round", "Participant 1", "Score 1", "Participant 2", "Score 2",
         "Winner", "Status", "Table/Court", "Scheduled At", "Finished At"],
        [
            [m["event_name"], m["sport_key"], (m["stage"] or "").replace("_", " ").title(), m["round"],
             m["participant_1"], m["score_1"], m["participant_2"], m["score_2"],
             m["winner"] or "—", (m["status"] or "").title(), m["table_number"] or "—",
             m["scheduled_at"] or "—", m["finished_at"] or "—"]
            for m in match_rows
        ],
        number_cols={3, 5, 7},
    )

    # ── Sheet 4: Participants ────────────────────────────────────
    ws4 = wb.create_sheet("Participants")
    _write_sheet(
        ws4,
        ["Event", "Sport", "Name", "Type", "Age", "Gender", "Seed", "Group", "Payment Status"],
        [
            [p["event_name"], p["sport_key"], p["name"], p["type"], p["age"] or "—", p["gender"] or "—",
             p["seed"] if p["seed"] is not None else "—", p["group_name"] or "—", (p["payment_status"] or "").replace("_", " ").title()]
            for p in participant_rows
        ],
        number_cols={4, 6},
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_filename(tournament_name_slug: str, on: date | None = None) -> str:
    d = on or date.today()
    return f"thescoreboard_{tournament_name_slug}_{d.isoformat()}.xlsx"

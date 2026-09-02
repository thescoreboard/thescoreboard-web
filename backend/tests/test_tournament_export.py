"""
Excel export — GET /tournaments/{id}/export.

Covers: authorization (staff+ allowed, outsider blocked, nonexistent 404),
response headers/content-type, workbook validity, expected data present,
and that exporting one tournament never leaks another tournament's data.
"""
import io
import os

os.environ["DATABASE_URL"] = "sqlite://"

from datetime import datetime, timezone

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import get_db
from app.main import app
from app.models.organization import Organization, OrgMember
from app.models.tournament import Tournament
from app.models.tournament_member import TournamentMember
from app.models.event import Event
from app.models.match import Match, MatchParticipant
from app.models.player import Player, Team
from app.models.group import EventParticipant
from app.models.user import User
from app.utils.auth import get_current_user


def _build_tournament(db, org, name, slug):
    t = Tournament(org_id=org.org_id, name=name, slug=slug, venue="Community Hall", city="Pune")
    db.add(t); db.flush()

    ev = Event(tournament_id=t.tournament_id, name="TT Singles",
               sport_key="table_tennis", format="round_robin", participant_type="individual")
    db.add(ev); db.flush()

    p1 = Player(name=f"{name} Player One", org_id=org.org_id, age=22, gender="male")
    p2 = Player(name=f"{name} Player Two", org_id=org.org_id, age=24, gender="female")
    db.add_all([p1, p2]); db.flush()

    db.add_all([
        EventParticipant(event_id=ev.event_id, player_id=p1.player_id, payment_status="not_required"),
        EventParticipant(event_id=ev.event_id, player_id=p2.player_id, payment_status="not_required"),
    ])

    m = Match(event_id=ev.event_id, round=1, stage="round_robin", status="done",
              finished_at=datetime.now(timezone.utc))
    db.add(m); db.flush()
    db.add_all([
        MatchParticipant(match_id=m.match_id, player_id=p1.player_id, position=1, score=3, is_winner=True),
        MatchParticipant(match_id=m.match_id, player_id=p2.player_id, position=2, score=1, is_winner=False),
    ])
    db.commit()
    return t, ev, p1, p2, m


def _setup(db):
    org = Organization(name="Club", slug="club")
    db.add(org); db.flush()

    owner    = User(email="owner@x.com", name="Owner", plan="free")
    staff    = User(email="staff@x.com", name="Staff", plan="free")
    outsider = User(email="outsider@x.com", name="Outsider", plan="free")
    db.add_all([owner, staff, outsider]); db.flush()
    db.add(OrgMember(org_id=org.org_id, user_id=owner.user_id, role="admin"))

    t, ev, p1, p2, m = _build_tournament(db, org, "Summer Open", "summer-open")
    db.add(TournamentMember(tournament_id=t.tournament_id, user_id=staff.user_id, role="staff"))
    db.commit()

    current = {"user": owner}
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: current["user"]

    users = {"owner": owner, "staff": staff, "outsider": outsider}
    return t, ev, users, current, org, TestClient(app)


def _as(current, users, who):
    current["user"] = users[who]


def test_owner_can_export(db):
    t, ev, users, current, org, client = _setup(db)
    try:
        r = client.get(f"/api/orgs/tournaments/{t.tournament_id}/export")
        assert r.status_code == 200, r.text
        assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in r.headers["content-disposition"]
        assert "thescoreboard_summer-open_" in r.headers["content-disposition"]
        assert r.headers["content-disposition"].endswith('.xlsx"')

        wb = load_workbook(io.BytesIO(r.content))
        assert "Summary" in wb.sheetnames
        assert "Standings" in wb.sheetnames  # round_robin event → standings present
        assert "Match Results" in wb.sheetnames
        assert "Participants" in wb.sheetnames
    finally:
        app.dependency_overrides.clear()


def test_staff_can_export(db):
    t, ev, users, current, org, client = _setup(db)
    try:
        _as(current, users, "staff")
        r = client.get(f"/api/orgs/tournaments/{t.tournament_id}/export")
        assert r.status_code == 200, r.text
    finally:
        app.dependency_overrides.clear()


def test_outsider_cannot_export(db):
    t, ev, users, current, org, client = _setup(db)
    try:
        _as(current, users, "outsider")
        r = client.get(f"/api/orgs/tournaments/{t.tournament_id}/export")
        assert r.status_code == 403, r.text
    finally:
        app.dependency_overrides.clear()


def test_nonexistent_tournament_404(db):
    t, ev, users, current, org, client = _setup(db)
    try:
        r = client.get("/api/orgs/tournaments/999999/export")
        assert r.status_code == 404, r.text
    finally:
        app.dependency_overrides.clear()


def test_export_contains_expected_data(db):
    t, ev, users, current, org, client = _setup(db)
    try:
        r = client.get(f"/api/orgs/tournaments/{t.tournament_id}/export")
        wb = load_workbook(io.BytesIO(r.content))

        summary = wb["Summary"]
        assert summary["A1"].value == "Summer Open"

        participants_sheet = wb["Participants"]
        names = {row[2] for row in participants_sheet.iter_rows(min_row=2, values_only=True)}
        assert "Summer Open Player One" in names
        assert "Summer Open Player Two" in names

        results_sheet = wb["Match Results"]
        rows = list(results_sheet.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][4] == "Summer Open Player One"  # participant 1
        assert rows[0][8] == "Summer Open Player One"  # winner

        standings_sheet = wb["Standings"]
        standings_rows = list(standings_sheet.iter_rows(min_row=2, values_only=True))
        assert any(row[3] == "Summer Open Player One" and row[2] == 1 for row in standings_rows)
    finally:
        app.dependency_overrides.clear()


def test_export_does_not_leak_other_tournament_data(db):
    t, ev, users, current, org, client = _setup(db)
    try:
        # A second, unrelated tournament in the same org with distinct data
        t2, ev2, p1b, p2b, m2 = _build_tournament(db, org, "Winter Cup", "winter-cup")

        r = client.get(f"/api/orgs/tournaments/{t.tournament_id}/export")
        assert r.status_code == 200, r.text
        wb = load_workbook(io.BytesIO(r.content))

        participants_sheet = wb["Participants"]
        names = {row[2] for row in participants_sheet.iter_rows(min_row=2, values_only=True)}
        assert "Winter Cup Player One" not in names
        assert "Winter Cup Player Two" not in names

        summary = wb["Summary"]
        assert summary["A1"].value == "Summer Open"
    finally:
        app.dependency_overrides.clear()
